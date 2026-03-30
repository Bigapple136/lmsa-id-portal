"""
LMSA Student Template Generator — pure Python + openpyxl (no JSZip/ZIP manipulation).
Produces a clean .xlsx that Excel will open without errors.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

DARK_NAVY   = "0D1B2A"
MED_GREEN   = "1A3A1A"
EMERALD     = "00653C"
GOLD        = "C9A84C"
LIGHT_GRAY  = "F3F4F6"
WHITE       = "FFFFFF"
MED_GRAY    = "D1D5DB"
DARK_GRAY   = "6B7280"
WARN_BG     = "FFF8E1"
WARN_TEXT   = "92400E"

COLS = ["student_id","full_name","year_level","position","programme",
        "blood_type","student_email","emergency_contact_name","emergency_contact_phone"]

COL_META = {
    "student_id":             (22, "Required. Unique. Format: AMD-2024-0001"),
    "full_name":              (28, "Required. As on enrollment form."),
    "year_level":             (16, "Required. Select from dropdown."),
    "position":               (22, "Optional. e.g. Class Representative."),
    "programme":              (20, "QR code only. e.g. MBBS, Pharm.D"),
    "blood_type":             (12, "QR code only. Select from dropdown."),
    "student_email":          (28, "QR code only. Email address."),
    "emergency_contact_name": (28, "QR code only. Full name."),
    "emergency_contact_phone":(22, "QR code only. e.g. +231 770 000000"),
}

LABELS = {
    "student_id":             "* Student ID",
    "full_name":              "* Full Name",
    "year_level":             "* Year Level",
    "position":               "Position",
    "programme":              "Programme",
    "blood_type":             "Blood Type",
    "student_email":          "Student Email",
    "emergency_contact_name": "Emergency Contact Name",
    "emergency_contact_phone":"Emergency Contact Phone",
}

YEAR_OPTIONS  = ["1st Year","2nd Year","3rd Year","4th Year","5th Year","6th Year"]
BLOOD_OPTIONS = ["A+","A-","B+","B-","AB+","AB-","O+","O-"]

def thick(color):  return Side(style="medium", color=color)
def thin(color):   return Side(style="thin",   color=color)
def s_fill(c):     return PatternFill("solid", fgColor=c)
def s_border(left=None, right=None, top=None, bottom=None):
    return Border(left=left, right=right, top=top, bottom=bottom)

QR_KEYS = {"programme","blood_type","student_email",
           "emergency_contact_name","emergency_contact_phone"}

def apply_title(ws, cell, text):
    cell.value = text
    cell.font  = Font(name="Calibri", size=14, bold=True, color=GOLD)
    cell.fill  = s_fill(DARK_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_subtitle(ws, cell, text):
    cell.value = text
    cell.font  = Font(name="Calibri", size=11, color=WHITE)
    cell.fill  = s_fill(MED_GREEN)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_label(cell, text):
    cell.value = text
    cell.font  = Font(name="Calibri", size=11, bold=True, color=GOLD)
    cell.fill  = s_fill(LIGHT_GRAY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = s_border(left=thin(MED_GRAY), top=thin(MED_GRAY), bottom=thin(MED_GRAY))

def apply_input(cell):
    cell.fill  = s_fill(WHITE)
    cell.font  = Font(name="Calibri", size=11, color="1A1A1A")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = s_border(
        left=thick(DARK_NAVY), right=thick(DARK_NAVY),
        top=thin(MED_GRAY), bottom=thin(MED_GRAY)
    )

def apply_note(cell, text):
    cell.value = text
    cell.font  = Font(name="Calibri", size=10, italic=True, color=DARK_GRAY)
    cell.fill  = s_fill("F9FAFB")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = s_border(right=thin(MED_GRAY), top=thin(MED_GRAY), bottom=thin(MED_GRAY))

def build_sheet1(wb):
    ws = wb.active
    ws.title = "Sheet1"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 52
    ws.column_dimensions["E"].width = 3

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:E1")
    apply_title(ws, ws["A1"], "LMSA Student ID Card Portal")

    ws.row_dimensions[2].height = 24
    ws.merge_cells("A2:E2")
    apply_subtitle(ws, ws["A2"],
        "Single Student Entry Form  \u2192  Fill in below, then use the Developer tab to add buttons.")

    ws.row_dimensions[3].height = 8

    for i, key in enumerate(COLS):
        row = 4 + i
        ws.row_dimensions[row].height = 22
        apply_label(ws[f"B{row}"], LABELS[key])
        apply_input(ws[f"C{row}"])
        apply_note(ws[f"D{row}"], COL_META[key][1])

    ws.row_dimensions[13].height = 12

    ws.row_dimensions[14].height = 32
    s = ws["B14"]
    s.value = "Save Record"
    s.font  = Font(name="Calibri", size=12, bold=True, color=WHITE)
    s.fill  = s_fill(EMERALD)
    s.alignment = Alignment(horizontal="center", vertical="center")

    n = ws["C14"]
    n.value = "New Record"
    n.font  = Font(name="Calibri", size=12, bold=True, color=EMERALD)
    n.fill  = s_fill(WHITE)
    n.border = s_border(left=thick(EMERALD), right=thick(EMERALD),
                        top=thick(EMERALD), bottom=thick(EMERALD))
    n.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[15].height = 24
    ws.merge_cells("A15:E15")
    note = ws["A15"]
    note.value = ("\u26A0  Macros must be enabled. Enable Developer tab > Visual Basic > "
                  "paste VBA code. See Instructions sheet.")
    note.font  = Font(name="Calibri", size=9, color=WARN_TEXT)
    note.fill  = s_fill(WARN_BG)
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    dv_year = DataValidation(
        type="list",
        formula1='"' + ','.join(YEAR_OPTIONS) + '"',
        allow_blank=True,
        showDropDown=False,
        showInputMessage=True,
        promptTitle="Year Level",
        prompt="Select from the dropdown."
    )
    ws.add_data_validation(dv_year)
    dv_year.add("C6")

    dv_blood = DataValidation(
        type="list",
        formula1='"' + ','.join(BLOOD_OPTIONS) + '"',
        allow_blank=True,
        showDropDown=False,
        showInputMessage=True,
        promptTitle="Blood Type",
        prompt="Select from the dropdown."
    )
    ws.add_data_validation(dv_blood)
    dv_blood.add("C9")

    return ws

def build_students(wb):
    ws = wb.create_sheet("Students")

    for i, key in enumerate(COLS):
        ws.column_dimensions[get_column_letter(i + 1)].width = COL_META[key][0]

    ws.row_dimensions[1].height = 22
    for i, key in enumerate(COLS):
        cell = ws.cell(row=1, column=i + 1)
        cell.value = key
        is_qr = key in QR_KEYS
        cell.font  = Font(name="Calibri", size=11, bold=True,
                           color="88CC88" if is_qr else GOLD)
        cell.fill  = s_fill(MED_GREEN if is_qr else DARK_NAVY)
        cell.alignment = Alignment(vertical="center", horizontal="left")

    sample = ["AMD-2024-0001","Josephine K. Freeman","3rd Year",
              "Class Representative","MBBS","O+",
              "josephine@email.com","Mary Freeman","+231 770 000000"]
    ws.row_dimensions[2].height = 20
    for i, val in enumerate(sample):
        cell = ws.cell(row=2, column=i + 1)
        cell.value = val
        cell.font  = Font(name="Calibri", size=10, italic=True, color="888780")
        cell.fill  = s_fill(LIGHT_GRAY)

    for r in range(3, 21):
        ws.row_dimensions[r].height = 18
        bg = "F9FAFB" if r % 2 == 0 else WHITE
        for i in range(len(COLS)):
            ws.cell(row=r, column=i + 1).fill = s_fill(bg)

    dv_year = DataValidation(
        type="list",
        formula1='"' + ','.join(YEAR_OPTIONS) + '"',
        allow_blank=True,
        showDropDown=False,
        showInputMessage=True,
        promptTitle="Year Level",
        prompt="Select from the dropdown."
    )
    ws.add_data_validation(dv_year)
    dv_year.add("C3:C1000")

    return ws

VBA_CODE = """' ============================================================
' LMSA Portal - Student Form Macros
' Copy everything between the === lines into the VBA Editor
' ============================================================

Sub SaveRecord()
    Dim wsForm As Worksheet, wsData As Worksheet
    Dim lastRow As Long, i As Integer
    Dim sid As String, r As Long

    Set wsForm = ThisWorkbook.Worksheets("Sheet1")
    Set wsData = ThisWorkbook.Worksheets("Students")

    If Trim(wsForm.Range("C4").Value) = "" Then
        MsgBox "Student ID is required!", vbCritical, "Missing Field"
        wsForm.Range("C4").Select
        Exit Sub
    End If
    If Trim(wsForm.Range("C5").Value) = "" Then
        MsgBox "Full Name is required!", vbCritical, "Missing Field"
        wsForm.Range("C5").Select
        Exit Sub
    End If
    If Trim(wsForm.Range("C6").Value) = "" Then
        MsgBox "Year Level is required!", vbCritical, "Missing Field"
        wsForm.Range("C6").Select
        Exit Sub
    End If

    sid = Trim(wsForm.Range("C4").Value)
    For r = 2 To wsData.Cells(wsData.Rows.Count, 1).End(xlUp).Row
        If Trim(wsData.Cells(r, 1).Value) = sid Then
            MsgBox "Student ID '" & sid & "' already exists in the Students sheet!", vbExclamation, "Duplicate Entry"
            Exit Sub
        End If
    Next r

    lastRow = wsData.Cells(wsData.Rows.Count, 1).End(xlUp).Row + 1
    If lastRow < 3 Then lastRow = 3

    Dim formRows: formRows = Array(4, 5, 6, 7, 8, 9, 10, 11, 12)
    For i = 0 To 8
        wsData.Cells(lastRow, i + 1).Value = Trim(wsForm.Cells(formRows(i), 3).Value)
    Next i

    MsgBox "Record saved!" & vbCrLf & vbCrLf & _
           "Student: " & wsForm.Range("C5").Value & vbCrLf & _
           "ID: " & wsForm.Range("C4").Value & vbCrLf & vbCrLf & _
           "Row " & lastRow & " in Students sheet.", _
           vbInformation, "Saved!"
End Sub

Sub NewRecord()
    Dim wsForm As Worksheet, i As Integer
    Set wsForm = ThisWorkbook.Worksheets("Sheet1")
    Dim formRows: formRows = Array(4, 5, 6, 7, 8, 9, 10, 11, 12)
    For i = 0 To 8
        wsForm.Cells(formRows(i), 3).Value = ""
    Next i
    wsForm.Range("C4").Select
End Sub"""

def build_instructions(wb):
    ws = wb.create_sheet("Instructions")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 90

    def hdr_row(row, height, texts):
        ws.row_dimensions[row].height = height
        for c, t in enumerate(texts):
            cell = ws.cell(row=row, column=c + 1)
            cell.value = t
            cell.font  = Font(name="Calibri", size=12, bold=True, color=WHITE)
            cell.fill  = s_fill(DARK_NAVY)
            cell.alignment = Alignment(vertical="center", horizontal="left")

    def data_row(row, height, texts):
        ws.row_dimensions[row].height = height
        for c, t in enumerate(texts):
            cell = ws.cell(row=row, column=c + 1)
            cell.value = t
            cell.font  = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    r = 1
    hdr_row(r, 30, ["LMSA Student ID Card Portal \u2014 Setup Guide","","",""]); r += 1
    ws.row_dimensions[r].height = 8; r += 1

    hdr_row(r, 22, ["SECTION 1: HOW TO SET UP THE MACROS","","",""]); r += 1
    ws.row_dimensions[r].height = 8; r += 1
    for step in [
        ("Step 1","Enable the Developer Tab",
         'File > Options > Customize Ribbon > check "Developer" in the right panel > OK'),
        ("Step 2","Open the VBA Editor",
         "Press Alt+F11 (or Developer tab > Visual Basic)"),
        ("Step 3","Insert a New Module",
         "In the VBA Editor: Insert > Module. A new code window opens."),
        ("Step 4","Paste the VBA Code",
         'Select ALL the VBA code from the "VBA SOURCE CODE" section below,\ncopy it, paste into the module window, then close the VBA Editor.'),
        ("Step 5","Add the Buttons (optional)",
         'Sheet1 > Developer tab > Insert > Button (Form Control).\nDraw a button, assign "SaveRecord" macro. Repeat for "NewRecord".\nTip: right-click each button to edit its text label.'),
        ("Step 6","Save as Macro-Enabled",
         'File > Save As > Browse > Save as type: Excel Macro-Enabled Workbook (*.xlsm)'),
    ]:
        data_row(r, 30, step); r += 1
    ws.row_dimensions[r].height = 8; r += 1

    hdr_row(r, 22, ["SECTION 2: HOW TO USE THE FORM","","",""]); r += 1
    ws.row_dimensions[r].height = 8; r += 1
    for item in [
        ("1","Go to the Student Form sheet","Sheet1 is the entry form."),
        ("2","Fill in the required fields","Fields marked with * are required (Student ID, Full Name, Year Level)."),
        ("3","Click Save Record","This appends your entry to the Students sheet."),
        ("4","Click New Record","This clears the form so you can enter the next student."),
        ("5","Year Level must match exactly",", ".join(YEAR_OPTIONS)),
    ]:
        data_row(r, 18, item); r += 1
    ws.row_dimensions[r].height = 8; r += 1

    hdr_row(r, 22, ["SECTION 3: COLUMN REFERENCE","","",""]); r += 1
    data_row(r, 18, ["Column","Required","Type","Notes"]); r += 1
    for key in COLS:
        data_row(r, 18, [
            key,
            "Yes" if key in ("student_id","full_name","year_level") else "No",
            "QR code only" if key in QR_KEYS else "Card face",
            COL_META[key][1],
        ]); r += 1
    ws.row_dimensions[r].height = 8; r += 1

    hdr_row(r, 22, ["SECTION 4: VBA SOURCE CODE","","",""]); r += 1
    data_row(r, 30, ["Copy everything below and paste into the VBA module:","","",""]); r += 1
    ws.row_dimensions[r].height = 500
    ws.merge_cells(f"A{r}:D{r}")
    cell = ws.cell(row=r, column=1)
    cell.value = VBA_CODE
    cell.font  = Font(name="Courier New", size=9, color="1A1A1A")
    cell.fill  = s_fill("F3F4F6")
    cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    r += 1
    ws.row_dimensions[r].height = 8; r += 1

    hdr_row(r, 22, ["SECTION 5: QR CODE FIELDS","","",""]); r += 1
    data_row(r, 40, [
        "Green-header columns in the Students sheet are encoded in the QR code only \u2014 "
        "they do not appear on the printed card face.",
        "","",""
    ])

    return ws

def generate(output_path):
    wb = Workbook()
    build_sheet1(wb)
    build_students(wb)
    build_instructions(wb)
    wb.save(output_path)
    size = os.path.getsize(output_path)
    print(f"Generated: {output_path}  ({size:,} bytes)")

if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    templates_dir = os.path.join(script_dir, "..", "templates")
    os.makedirs(templates_dir, exist_ok=True)
    out = os.path.join(templates_dir, "student_form_template.xlsx")
    generate(out)
